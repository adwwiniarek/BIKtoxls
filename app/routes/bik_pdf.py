# app/routes/bik_pdf.py
from __future__ import annotations
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from io import BytesIO
import httpx, fitz, re, os
from typing import Any, Dict, List, Optional, Tuple
from ..notion_client import NotionClient, NotionError

router = APIRouter()

# ---------- Pobieranie PDF ----------
def _fetch_url(url: str) -> bytes:
    try:
        with httpx.Client(timeout=60) as cli:
            r = cli.get(url, follow_redirects=True)
            r.raise_for_status()
            return r.content
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Nie można pobrać PDF: {e}")

def _bik_pdfs_from_notion(page_id: str) -> List[Tuple[str, bytes]]:
    """
    Zwraca LISTĘ (nazwa, bytes) wszystkich PDF-ów z właściwości 'Raporty BIK'.
    Jeśli brak PDF-ów -> 404.
    """
    client = NotionClient()
    try:
        page = client.get_page(page_id)
    except NotionError as e:
        raise HTTPException(status_code=502, detail=f"Notion error: {e}")

    props = page.get("properties", {}) or {}
    files = (props.get("Raporty BIK") or {}).get("files", []) or []
    out: List[Tuple[str, bytes]] = []

    for f in files:
        nm = f.get("name") or "raport.pdf"
        if f.get("type") == "file":
            url = (f.get("file") or {}).get("url")
        else:
            url = (f.get("external") or {}).get("url")
        if url and url.lower().endswith(".pdf"):
            out.append((nm, _fetch_url(url)))

    if not out:
        raise HTTPException(status_code=404, detail="Nie znaleziono PDF-ów w 'Raporty BIK'.")
    return out

def _text_from_pdf(pdf_bytes: bytes) -> str:
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        parts: List[str] = []
        for page in doc:
            parts.append(page.get_text("text"))
        return "\n".join(parts)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"PDF nieczytelny: {e}")

# ---------- Heurystyczny parser BIK ----------
RE_ROW = re.compile(
    r"(?P<Kredytodawca>.+?)\s+"
    r"(?P<Rodzaj>Kredyt|Pożyczka|Karta|Limit|Leasing|Faktoring).*?"
    r"Nr\s*umowy[:\s]*(?P<NrUmowy>[\w/-]+).*?"
    r"Data\s*uruchomienia[:\s]*(?P<DataUruchomienia>\d{4}-\d{2}-\d{2}|\d{2}\.\d{2}\.\d{4}).*?"
    r"Saldo[:\s]*(?P<Saldo>[-\d\s,.]+)\s*PLN.*?"
    r"Zaległość[:\s]*(?P<Zaleglosc>[-\d\s,.]+)\s*PLN.*?"
    r"Status[:\s]*(?P<Status>Aktywny|Wypowiedziany|Zamknięty|Windykacja)",
    re.IGNORECASE | re.DOTALL
)

def _num(txt: str) -> float:
    if txt is None:
        return 0.0
    t = txt.replace(" ", "").replace("\u00a0","").replace(",",".")
    try:
        return float(re.sub(r"[^0-9.\-]", "", t))
    except:
        return 0.0

def parse_bik(text: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    chunks = re.split(r"(?:\n\s*){2,}", text)  # proste rozbicie na sekcje
    for ch in chunks:
        m = RE_ROW.search(ch)
        if not m:
            continue
        gd = m.groupdict()
        rows.append({
            "Kredytodawca": (gd.get("Kredytodawca") or "").strip(),
            "Rodzaj": (gd.get("Rodzaj") or "").capitalize(),
            "Nr umowy": (gd.get("NrUmowy") or "").strip(),
            "Data uruchomienia": (gd.get("DataUruchomienia") or "").replace(".", "-"),
            "Saldo (PLN)": _num(gd.get("Saldo")),
            "Zaległość (PLN)": _num(gd.get("Zaleglosc")),
            "Status": (gd.get("Status") or "").capitalize(),
        })
    return rows

# ---------- Wspólna budowa XLS ----------
def _build_xls(rows: List[Dict[str, Any]], filename: str = "BIK_z_raportu.xlsx") -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Zobowiązania (BIK)"
    headers = ["Kredytodawca", "Rodzaj", "Nr umowy", "Data uruchomienia", "Saldo (PLN)", "Zaległość (PLN)", "Status"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    rr = 2
    for row in rows:
        ws.cell(row=rr, column=1, value=row.get("Kredytodawca",""))
        ws.cell(row=rr, column=2, value=row.get("Rodzaj",""))
        ws.cell(row=rr, column=3, value=row.get("Nr umowy",""))
        ws.cell(row=rr, column=4, value=row.get("Data uruchomienia",""))
        ws.cell(row=rr, column=5, value=row.get("Saldo (PLN)",0.0))
        ws.cell(row=rr, column=6, value=row.get("Zaległość (PLN)",0.0))
        ws.cell(row=rr, column=7, value=row.get("Status",""))
        rr += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 16
    for r in ws.iter_rows(min_row=2, min_col=1, max_col=7, max_row=rr-1):
        for cell in r:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\""}
    )

# ---------- Endpoint: PDF -> XLSX (łączy wiele PDF-ów) ----------
@router.post("/bik/pdf-to-xls")
async def bik_pdf_to_xls(
    pdf_url: Optional[str] = Form(None, description="URL do jednego PDF raportu BIK"),
    notion_page_id: Optional[str] = Form(None, description="Notion page_id — zbierze WSZYSTKIE PDF-y z 'Raporty BIK'"),
    file: Optional[UploadFile] = File(None, description="Lub bezpośredni upload jednego PDF")
):
    rows_all: List[Dict[str, Any]] = []

    # 1) Notion: WEŹ WSZYSTKIE PDF-y z właściwości „Raporty BIK” i połącz
    if notion_page_id:
        pdfs = _bik_pdfs_from_notion(notion_page_id)  # lista (nazwa, bytes)
        for _, pdf_bytes in pdfs:
            text = _text_from_pdf(pdf_bytes)
            rows_all.extend(parse_bik(text))
        fname = "BIK_z_raportow.pdf.xlsx" if len(pdfs) > 1 else "BIK_z_raportu.pdf.xlsx"
        return _build_xls(rows_all, filename=fname)

    # 2) Upload jednego PDF-a
    if file is not None:
        pdf_bytes = await file.read()
        rows_all = parse_bik(_text_from_pdf(pdf_bytes))
        return _build_xls(rows_all, filename="BIK_z_raportu.pdf.xlsx")

    # 3) Jeden PDF z URL
    if pdf_url:
        pdf_bytes = _fetch_url(pdf_url)
        rows_all = parse_bik(_text_from_pdf(pdf_bytes))
        return _build_xls(rows_all, filename="BIK_z_raportu.pdf.xlsx")

    raise HTTPException(status_code=422, detail="Podaj PDF (plik), albo pdf_url, albo notion_page_id.")

# ---------- KOMPATYBILNOŚĆ: stary URL -> łączy wiele PDF-ów ----------
@router.get("/notion/poll-one")
def notion_poll_one_compat(page_id: str = Query(...), x_key: str | None = Query(None)):
    """
    GET /notion/poll-one?page_id=...&x_key=...
    Zwraca JEDEN XLS, który łączy WSZYSTKIE PDF-y z 'Raporty BIK' dla danej strony Notion.
    (Tylko dane z PDF; bez wzbogacania po NIP.)
    """
    required = os.getenv("NOTION_X_KEY")
    if required and x_key != required:
        raise HTTPException(status_code=403, detail="Forbidden")

    rows_all: List[Dict[str, Any]] = []
    pdfs = _bik_pdfs_from_notion(page_id)  # lista (nazwa, bytes)
    for _, pdf_bytes in pdfs:
        text = _text_from_pdf(pdf_bytes)
        rows_all.extend(parse_bik(text))

    fname = "BIK_z_raportow.pdf.xlsx" if len(pdfs) > 1 else "BIK_z_raportu.pdf.xlsx"
    return _build_xls(rows_all, filename=fname)
